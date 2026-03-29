"""Excel export utility for reports."""
import os
import tempfile
from datetime import date
from decimal import Decimal

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

HEADER_FONT = Font(bold=True, size=11)
HEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
CURRENCY_FORMAT = '#,##0.00'
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'),
)


def _style_header(ws, row, cols):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='center')


def _auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = None
        for cell in col:
            if hasattr(cell, 'column_letter'):
                col_letter = cell.column_letter
            if cell.value and not isinstance(cell, type(None)):
                try:
                    max_len = max(max_len, len(str(cell.value)))
                except Exception:
                    pass
        if col_letter:
            ws.column_dimensions[col_letter].width = min(max_len + 3, 40)


def generate_totals_excel(
    title: str,
    period_start: date,
    period_end: date,
    totals_data: dict,
    currency: str,
    lang_labels: dict,
) -> str:
    """Generate totals-only Excel (no worker breakdown).
    
    totals_data: {type_name: {unit, quantity, total}}
    Returns: file path
    """
    wb = Workbook()
    ws = wb.active
    ws.title = lang_labels.get("summary", "Summary")

    ws.merge_cells('A1:D1')
    ws['A1'] = title
    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'] = f"{period_start.strftime('%d.%m.%y')} — {period_end.strftime('%d.%m.%y')}"
    ws['A2'].font = Font(size=11, italic=True)

    row = 4
    headers = [
        lang_labels.get("work_type", "Work type"),
        lang_labels.get("unit", "Unit"),
        lang_labels.get("quantity", "Quantity"),
        f"{lang_labels.get('total', 'Total')} ({currency})",
    ]
    for i, h in enumerate(headers, 1):
        ws.cell(row=row, column=i, value=h)
    _style_header(ws, row, len(headers))

    row += 1
    grand_total = Decimal("0")

    for tname, tdata in totals_data.items():
        ws.cell(row=row, column=1, value=tname)
        ws.cell(row=row, column=2, value=tdata["unit"])
        ws.cell(row=row, column=3, value=float(tdata["quantity"]))
        ws.cell(row=row, column=4, value=float(tdata["total"]))
        ws.cell(row=row, column=4).number_format = CURRENCY_FORMAT
        for c in range(1, 5):
            ws.cell(row=row, column=c).border = THIN_BORDER
        grand_total += tdata["total"]
        row += 1

    row += 1
    ws.cell(row=row, column=3, value=lang_labels.get("grand_total", "TOTAL"))
    ws.cell(row=row, column=3).font = Font(bold=True, size=12)
    ws.cell(row=row, column=4, value=float(grand_total))
    ws.cell(row=row, column=4).font = Font(bold=True, size=12)
    ws.cell(row=row, column=4).number_format = CURRENCY_FORMAT

    _auto_width(ws)

    fd, path = tempfile.mkstemp(suffix=".xlsx")
    os.close(fd)
    wb.save(path)
    return path


def generate_summary_excel(
    title: str,
    period_start: date,
    period_end: date,
    worker_data: dict,
    currency: str,
    lang_labels: dict,
) -> str:
    """Generate summary Excel file.
    
    worker_data: {worker_name: {types: {type_name: {unit, quantity, total}}, total}}
    Returns: file path
    """
    wb = Workbook()
    ws = wb.active
    ws.title = lang_labels.get("summary", "Summary")

    # Title
    ws.merge_cells('A1:E1')
    ws['A1'] = title
    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'] = f"{period_start.strftime('%d.%m.%y')} — {period_end.strftime('%d.%m.%y')}"
    ws['A2'].font = Font(size=11, italic=True)

    # Headers
    row = 4
    headers = [
        lang_labels.get("worker", "Worker"),
        lang_labels.get("work_type", "Work type"),
        lang_labels.get("unit", "Unit"),
        lang_labels.get("quantity", "Quantity"),
        f"{lang_labels.get('total', 'Total')} ({currency})",
    ]
    for i, h in enumerate(headers, 1):
        ws.cell(row=row, column=i, value=h)
    _style_header(ws, row, len(headers))

    row += 1
    grand_total = Decimal("0")

    for wname, data in worker_data.items():
        first_row = row
        for tname, tdata in data["types"].items():
            ws.cell(row=row, column=1, value=wname if row == first_row else "")
            ws.cell(row=row, column=2, value=tname)
            ws.cell(row=row, column=3, value=tdata["unit"])
            ws.cell(row=row, column=4, value=float(tdata["quantity"]))
            ws.cell(row=row, column=5, value=float(tdata["total"]))
            ws.cell(row=row, column=5).number_format = CURRENCY_FORMAT
            for c in range(1, 6):
                ws.cell(row=row, column=c).border = THIN_BORDER
            row += 1

        # Worker subtotal
        ws.cell(row=row, column=1, value="")
        ws.cell(row=row, column=4, value=lang_labels.get("subtotal", "Subtotal"))
        ws.cell(row=row, column=4).font = HEADER_FONT
        ws.cell(row=row, column=5, value=float(data["total"]))
        ws.cell(row=row, column=5).font = HEADER_FONT
        ws.cell(row=row, column=5).number_format = CURRENCY_FORMAT
        grand_total += data["total"]
        row += 1

    # Grand total
    row += 1
    ws.cell(row=row, column=4, value=lang_labels.get("grand_total", "TOTAL"))
    ws.cell(row=row, column=4).font = Font(bold=True, size=12)
    ws.cell(row=row, column=5, value=float(grand_total))
    ws.cell(row=row, column=5).font = Font(bold=True, size=12)
    ws.cell(row=row, column=5).number_format = CURRENCY_FORMAT

    _auto_width(ws)

    fd, path = tempfile.mkstemp(suffix=".xlsx")
    os.close(fd)
    wb.save(path)
    return path


def generate_project_excel(
    title: str,
    period_start: date,
    period_end: date,
    project_data: dict,
    currency: str,
    lang_labels: dict,
) -> str:
    """Generate project report Excel.
    
    project_data: {project_name: {building_name|None: {elem_name|None: {worker: {types: {name: {unit,qty,total}}}}}}}
    Returns: file path
    """
    wb = Workbook()
    ws = wb.active
    ws.title = lang_labels.get("by_project", "By project")

    ws.merge_cells('A1:F1')
    ws['A1'] = title
    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'] = f"{period_start.strftime('%d.%m.%y')} — {period_end.strftime('%d.%m.%y')}"

    row = 4
    headers = [
        lang_labels.get("project", "Project"),
        lang_labels.get("worker", "Worker"),
        lang_labels.get("work_type", "Work type"),
        lang_labels.get("unit", "Unit"),
        lang_labels.get("quantity", "Quantity"),
        f"{lang_labels.get('total', 'Total')} ({currency})",
    ]
    for i, h in enumerate(headers, 1):
        ws.cell(row=row, column=i, value=h)
    _style_header(ws, row, len(headers))

    row += 1
    grand_total = Decimal("0")

    for pname, buildings in project_data.items():
        project_first = True
        project_total = Decimal("0")

        for bld_name, elems in buildings.items():
            location = pname
            if bld_name:
                location += f" → {bld_name}"
            for elem_name, workers in elems.items():
                if elem_name:
                    location_full = f"{location} → {elem_name}"
                else:
                    location_full = location

                for wname, wdata in workers.items():
                    for tname, tdata in wdata["types"].items():
                        ws.cell(row=row, column=1, value=location_full if project_first else "")
                        ws.cell(row=row, column=2, value=wname)
                        ws.cell(row=row, column=3, value=tname)
                        ws.cell(row=row, column=4, value=tdata["unit"])
                        ws.cell(row=row, column=5, value=float(tdata["qty"]))
                        ws.cell(row=row, column=6, value=float(tdata["total"]))
                        ws.cell(row=row, column=6).number_format = CURRENCY_FORMAT
                        for c in range(1, 7):
                            ws.cell(row=row, column=c).border = THIN_BORDER
                        project_first = False
                        project_total += tdata["total"]
                        row += 1

        # Project subtotal
        ws.cell(row=row, column=5, value=f"{pname} —")
        ws.cell(row=row, column=5).font = HEADER_FONT
        ws.cell(row=row, column=6, value=float(project_total))
        ws.cell(row=row, column=6).font = HEADER_FONT
        ws.cell(row=row, column=6).number_format = CURRENCY_FORMAT
        grand_total += project_total
        row += 1

    row += 1
    ws.cell(row=row, column=5, value=lang_labels.get("grand_total", "TOTAL"))
    ws.cell(row=row, column=5).font = Font(bold=True, size=12)
    ws.cell(row=row, column=6, value=float(grand_total))
    ws.cell(row=row, column=6).font = Font(bold=True, size=12)
    ws.cell(row=row, column=6).number_format = CURRENCY_FORMAT

    _auto_width(ws)

    fd, path = tempfile.mkstemp(suffix=".xlsx")
    os.close(fd)
    wb.save(path)
    return path
